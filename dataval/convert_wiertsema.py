import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def _sanitize_filename(name: str) -> str:
    s = (name or "").replace("(", "").replace(")", "")
    s = re.sub(r'[<>:"/\\|?*]+', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.strip(" ._")
    return s or "series"


def _norm_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).lower().replace("(", "").replace(")", "")
    return re.sub(r"[^a-z0-9]+", "", s)


def _first_non_empty(values) -> str:
    for v in values:
        if v not in (None, ""):
            return str(v)
    return "sheet"


def _to_datetime(v, wb_epoch):
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        try:
            return from_excel(v, wb_epoch)
        except Exception:
            return None
    return None


def _detect_sensor_groups(title_row, header_row) -> list[dict]:
    sensor_positions = []
    for col_idx in range(1, len(title_row)):
        val = title_row[col_idx]
        if val is not None and str(val).strip() != "":
            sensor_positions.append((col_idx, str(val).strip()))

    n_headers = len(header_row) if header_row else 0

    if not sensor_positions:
        wn_idx = None
        for i in range(n_headers):
            nh = _norm_header(str(header_row[i]) if header_row[i] is not None else "")
            if ("waterniveau" in nh or "grondwaterstand" in nh) and ("mnap" in nh or "nap" in nh):
                wn_idx = i
                break
        return [{"sensor_id": "unknown", "wn_idx": wn_idx, "start_col": 0, "end_col": n_headers}]

    groups = []
    for idx, (start_col, sensor_id) in enumerate(sensor_positions):
        end_col = sensor_positions[idx + 1][0] if idx + 1 < len(sensor_positions) else n_headers

        wn_idx = None
        for i in range(start_col, end_col):
            if i >= n_headers:
                break
            nh = _norm_header(str(header_row[i]) if header_row[i] is not None else "")
            if ("waterniveau" in nh or "grondwaterstand" in nh) and ("mnap" in nh or "nap" in nh):
                wn_idx = i
                break

        groups.append({
            "sensor_id": sensor_id,
            "wn_idx": wn_idx,
            "start_col": start_col,
            "end_col": end_col,
        })

    return groups


def process_workbook(
    xlsx_path: Path,
    dataset_root: Path,
    origin_name: str | None = None,
) -> list[dict]:
    print(f"Bestand: {xlsx_path.name}")

    origin_stem = origin_name or xlsx_path.stem
    out_folder = dataset_root / origin_stem / "only_csv"
    out_folder.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    written = 0
    written_paths: set[Path] = set()
    summary_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            title_row = next(rows_iter)
        except StopIteration:
            continue
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        base_title = _first_non_empty(title_row)
        safe_title = _sanitize_filename(base_title)

        headers = [str(h) if h is not None else "" for h in header_row]
        norm_headers = [_norm_header(h) for h in headers]

        ts_idx = None
        for i, nh in enumerate(norm_headers):
            if nh in ("timestamp", "time", "datetime"):
                ts_idx = i
                break

        groups = _detect_sensor_groups(title_row, header_row)
        wl_groups = [g for g in groups if g["wn_idx"] is not None]

        if ts_idx is None or not wl_groups:
            print(f"  Tabblad: {sheet_name} -> skipped (missing Timestamp or Waterniveau/Grondwaterstand column)")
            continue

        data_rows = list(rows_iter)

        summary_row: dict = {
            "series_name": base_title,
            "nr_sensors": len(groups),
            "nr_series": len(wl_groups),
        }

        for sensor_num, group in enumerate(wl_groups, start=1):
            wn_idx = group["wn_idx"]
            records = []

            for row in data_rows:
                if row is None:
                    continue
                ts_val = row[ts_idx] if ts_idx < len(row) else None
                wn_val = row[wn_idx] if wn_idx < len(row) else None

                ts = _to_datetime(ts_val, wb.epoch)
                if ts is None:
                    try:
                        ts = pd.to_datetime(ts_val, dayfirst=True, errors="coerce")
                        if pd.isna(ts):
                            continue
                    except Exception:
                        continue

                wn = pd.to_numeric(wn_val, errors="coerce")
                if pd.isna(wn):
                    continue

                records.append((ts, wn))

            if not records:
                print(f"  Tabblad: {sheet_name} sensor {sensor_num}/{len(wl_groups)} ({group['sensor_id']}) -> no valid rows")
                continue

            df = pd.DataFrame(records, columns=["Time", "head"]).set_index("Time")
            df.sort_index(inplace=True)

            summary_row[f"start_{sensor_num}"] = df.index.min()

            if len(wl_groups) == 1:
                out_path = out_folder / f"{safe_title}.csv"
            else:
                out_path = out_folder / f"{safe_title}_sensor_{sensor_num}.csv"

            if out_path in written_paths:
                safe_sheet = _sanitize_filename(sheet_name)
                out_path = out_folder / f"{out_path.stem} [{safe_sheet}].csv"

            df.to_csv(out_path, index=True, index_label="Time")
            written_paths.add(out_path)
            print(f"  Tabblad: {sheet_name} sensor {sensor_num}/{len(wl_groups)} ({group['sensor_id']}) -> Saved {out_path.name} ({len(df)} rows)")
            written += 1

        summary_rows.append(summary_row)

    wb.close()
    print(f"Done. Saved {written} CSVs to {out_folder}\n")
    return summary_rows


def convert_wiertsema_batch(
    input_dir: Path,
    output_dir: Path,
    file_index: int | None = None,
) -> list[dict]:
    from ._utils import parse_stable_key

    xlsx_files = sorted([p for p in input_dir.iterdir() if p.suffix.lower() == ".xlsx"])
    if not xlsx_files:
        print(f"No Excel files found in {input_dir}")
        return []

    print(f"Found {len(xlsx_files)} Excel file(s) in {input_dir.name}:")
    for i, f in enumerate(xlsx_files):
        print(f"  {i}: {f.name}")

    if file_index is not None:
        if file_index < 0 or file_index >= len(xlsx_files):
            raise ValueError(f"file_index {file_index} out of range (0-{len(xlsx_files)-1})")
        xlsx_files = [xlsx_files[file_index]]

    all_summary_rows: list[dict] = []
    for xfile in xlsx_files:
        origin_name = parse_stable_key(xfile.name, "wiertsema")
        rows = process_workbook(xfile, output_dir, origin_name=origin_name)
        all_summary_rows.extend(rows)

    if all_summary_rows:
        summary_df = pd.DataFrame(all_summary_rows)
        start_cols = sorted(
            [c for c in summary_df.columns if c.startswith("start_")],
            key=lambda c: int(c.split("_")[1]),
        )
        ordered_cols = ["series_name", "nr_sensors", "nr_series"] + start_cols
        summary_df = summary_df.reindex(columns=ordered_cols)
        summary_path = output_dir / "series_summary.csv"
        summary_df.to_csv(summary_path, index=False)
        print(f"Summary saved to {summary_path} ({len(summary_df)} rows)")

    return all_summary_rows
